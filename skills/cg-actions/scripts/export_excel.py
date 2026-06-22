#!/usr/bin/env python3
"""
export_excel.py — export inventory.json to a filterable Excel workbook.

Requires: pip install openpyxl

Produces three sheets:
  Summary   — org-level counts and coverage metrics
  Actions   — one row per distinct action: status, versions in use, repos,
              workflows, usages, hardened ref, and any version-gap note
  Usages    — one row per occurrence: repo, workflow, line, current ref,
              hardened ref, status, note — fully filterable for large orgs

Usage:
    python3 export_excel.py inventory.json
    python3 export_excel.py inventory.json --out migration-report.xlsx
"""
import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print(
        "error: openpyxl is required for Excel export.\n"
        "Install it with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
GREEN_BG  = PatternFill("solid", fgColor="D6F4D6")
TEAL_BG   = PatternFill("solid", fgColor="D4ECF2")  # already hardened
PURPLE_BG = PatternFill("solid", fgColor="E6DCF2")  # mixed (per version)
YELLOW_BG = PatternFill("solid", fgColor="FFF3CD")
RED_BG    = PatternFill("solid", fgColor="FCE4D6")
GREY_BG   = PatternFill("solid", fgColor="F2F2F2")
HEADER_BG = PatternFill("solid", fgColor="1A3A4A")  # Chainguard dark blue

GREEN_FONT  = Font(color="276221", bold=False)
TEAL_FONT   = Font(color="1A4A5A", bold=False)
PURPLE_FONT = Font(color="5A3A7A", bold=False)
YELLOW_FONT = Font(color="7D5700", bold=False)
RED_FONT    = Font(color="9C2700", bold=False)
GREY_FONT   = Font(color="666666", bold=False)
HEADER_FONT = Font(color="FFFFFF", bold=True)

THIN = Side(style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _status_label(ha):
    if ha is True:
        return "available"
    if ha == "already-hardened":
        return "already hardened"
    if ha == "mixed":
        return "mixed (per version)"
    if ha == "version-gap":
        return "version gap"
    if ha is False:
        return "no equivalent"
    return "not checked"


def _status_style(ha):
    if ha is True:
        return GREEN_BG, GREEN_FONT
    if ha == "already-hardened":
        return TEAL_BG, TEAL_FONT
    if ha == "mixed":
        return PURPLE_BG, PURPLE_FONT
    if ha == "version-gap":
        return YELLOW_BG, YELLOW_FONT
    if ha is False:
        return RED_BG, RED_FONT
    return GREY_BG, GREY_FONT


def _versions_label(action):
    refs = sorted({(o.get("version") or o.get("ref")) for o in action["occurrences"]
                   if o.get("version") or o.get("ref")})
    if not refs:
        return ""
    if len(refs) <= 3:
        return ", ".join(refs)
    return ", ".join(refs[:3]) + f" +{len(refs) - 3} more"


def _apply_header_row(ws, row, columns):
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _cell(ws, row, col, value, fill=None, font=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fill:
        c.fill = fill
    if font:
        c.font = font
    return c


def _freeze_and_autofilter(ws, header_row=1):
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_summary_sheet(wb, inv):
    ws = wb.create_sheet("Summary")
    actions = inv.get("actions", [])
    total_usages = sum(len(a["occurrences"]) for a in actions)
    all_repos = {o["repo"] for a in actions for o in a["occurrences"]}
    all_wfs = {(o["repo"], o["workflow"]) for a in actions for o in a["occurrences"]}
    avail = [a for a in actions if a.get("hardened_available") is True]
    already = [a for a in actions if a.get("hardened_available") == "already-hardened"]
    mixed = [a for a in actions if a.get("hardened_available") == "mixed"]
    gaps = [a for a in actions if a.get("hardened_available") == "version-gap"]
    no_equiv = [a for a in actions if a.get("hardened_available") is False]
    unknown = [a for a in actions if a.get("hardened_available") is None]

    # Usage-level tally (honest when actions are mixed across versions).
    occ_status = [o.get("hardened_available") for a in actions for o in a["occurrences"]]
    u_avail = sum(1 for s in occ_status if s is True)
    u_gap = sum(1 for s in occ_status if s == "version-gap")

    rows = [
        ("Org", inv.get("org") or "(unknown)"),
        ("Version strategy", inv.get("version_strategy", "same-major")),
        ("Generated by", inv.get("generated_by", "")),
        (None, None),
        ("Distinct actions", len(actions)),
        ("Total usages", total_usages),
        ("Repos scanned", len(all_repos)),
        ("Workflows scanned", len(all_wfs)),
        (None, None),
        ("✅  Available — ready to swap", len(avail)),
        ("🛡️  Already hardened — no change needed", len(already)),
        ("◐  Mixed — status varies by version", len(mixed)),
        ("⚠️  Version gap — major jump required", len(gaps)),
        ("❌  No hardened equivalent", len(no_equiv)),
        ("?   Not yet checked", len(unknown)),
        (None, None),
        ("Usages swappable now", u_avail),
        ("Usages needing version review", u_gap),
    ]

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28

    for r, (label, value) in enumerate(rows, start=1):
        if label is None:
            continue
        la = ws.cell(row=r, column=1, value=label)
        va = ws.cell(row=r, column=2, value=value)
        la.border = THIN_BORDER
        va.border = THIN_BORDER
        la.alignment = Alignment(vertical="center")
        va.alignment = Alignment(vertical="center")
        if label.startswith("✅"):
            la.fill, la.font = GREEN_BG, GREEN_FONT
            va.fill = GREEN_BG
        elif label.startswith("🛡️"):
            la.fill, la.font = TEAL_BG, TEAL_FONT
            va.fill = TEAL_BG
        elif label.startswith("◐"):
            la.fill, la.font = PURPLE_BG, PURPLE_FONT
            va.fill = PURPLE_BG
        elif label.startswith("⚠️"):
            la.fill, la.font = YELLOW_BG, YELLOW_FONT
            va.fill = YELLOW_BG
        elif label.startswith("❌"):
            la.fill, la.font = RED_BG, RED_FONT
            va.fill = RED_BG


def build_actions_sheet(wb, inv):
    ws = wb.create_sheet("Actions")
    columns = [
        ("Action",            36),
        ("Status",            16),
        ("Versions in use",   18),
        ("Repos",              8),
        ("Workflows",         12),
        ("Usages",             9),
        ("Hardened ref",      48),
        ("Note",              52),
    ]
    _apply_header_row(ws, 1, columns)

    actions = inv.get("actions", [])
    for r, a in enumerate(sorted(actions, key=lambda x: -len(x["occurrences"])), start=2):
        ha = a.get("hardened_available")
        fill, font = _status_style(ha)
        repos = len({o["repo"] for o in a["occurrences"]})
        wfs = len({(o["repo"], o["workflow"]) for o in a["occurrences"]})
        usages = len(a["occurrences"])

        _cell(ws, r, 1, a["name"])
        _cell(ws, r, 2, _status_label(ha), fill=fill, font=font)
        _cell(ws, r, 3, _versions_label(a))
        _cell(ws, r, 4, repos)
        _cell(ws, r, 5, wfs)
        _cell(ws, r, 6, usages)
        _cell(ws, r, 7, a.get("hardened_ref") or "")
        _cell(ws, r, 8, a.get("hardened_note") or "", wrap=True)

    _freeze_and_autofilter(ws)

    tbl = Table(displayName="Actions", ref=f"A1:{get_column_letter(len(columns))}{len(actions)+1}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)


def build_usages_sheet(wb, inv):
    ws = wb.create_sheet("Usages")
    columns = [
        ("Action",        36),
        ("Repo",          30),
        ("Workflow",      40),
        ("Line",           7),
        ("Current ref",   14),
        ("Status",        16),
        ("Hardened ref",  48),
        ("Note",          52),
    ]
    _apply_header_row(ws, 1, columns)

    row_idx = 2
    for a in sorted(inv.get("actions", []), key=lambda x: x["name"]):
        for occ in a["occurrences"]:
            # Per-occurrence status/ref/note so mixed-version actions read accurately.
            ha = occ.get("hardened_available", a.get("hardened_available"))
            href = occ.get("hardened_ref", a.get("hardened_ref"))
            note = occ.get("hardened_note", a.get("hardened_note"))
            fill, font = _status_style(ha)
            _cell(ws, row_idx, 1, a["name"])
            _cell(ws, row_idx, 2, occ.get("repo") or "")
            _cell(ws, row_idx, 3, occ.get("workflow") or "")
            _cell(ws, row_idx, 4, occ.get("line"))
            _cell(ws, row_idx, 5, occ.get("version") or occ.get("ref") or "")
            _cell(ws, row_idx, 6, _status_label(ha), fill=fill, font=font)
            _cell(ws, row_idx, 7, href or "")
            _cell(ws, row_idx, 8, note or "", wrap=True)
            row_idx += 1

    _freeze_and_autofilter(ws)

    total_rows = row_idx - 1
    if total_rows > 1:
        tbl = Table(displayName="Usages", ref=f"A1:{get_column_letter(len(columns))}{total_rows}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("inventory", help="Path to inventory.json")
    ap.add_argument("--out", help="Output .xlsx path (default: <inventory-stem>.xlsx)")
    args = ap.parse_args(argv)

    with open(args.inventory, "r", encoding="utf-8") as f:
        inv = json.load(f)

    out_path = args.out or str(Path(args.inventory).with_suffix(".xlsx"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    build_summary_sheet(wb, inv)
    build_actions_sheet(wb, inv)
    build_usages_sheet(wb, inv)

    wb.save(out_path)
    action_count = len(inv.get("actions", []))
    usage_count = sum(len(a["occurrences"]) for a in inv.get("actions", []))
    print(f"Wrote {out_path}  ({action_count} actions, {usage_count} usages across 3 sheets)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
